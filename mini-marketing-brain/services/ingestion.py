import csv
from pathlib import Path

from services.storage import save_knowledge_entry


def ingest_file(file_path: str, org_slug: str, title: str, doc_id: str) -> int:
    """Parse a file and create knowledge entries. Returns number of entries created."""
    path = Path(file_path)
    suffix = path.suffix.lower()

    if suffix == ".csv":
        return _ingest_csv(file_path, org_slug, title, doc_id)
    elif suffix in (".xlsx", ".xls"):
        return _ingest_xlsx(file_path, org_slug, title, doc_id)
    elif suffix == ".txt":
        text = path.read_text(encoding="utf-8", errors="ignore")
        save_knowledge_entry(text, org_slug, title, doc_id)
        return 1
    else:
        # Store raw text representation for unsupported types
        text = f"[Uploaded file: {path.name}] (format {suffix} — text extraction not supported)"
        save_knowledge_entry(text, org_slug, title, doc_id)
        return 1


def ingest_text(text: str, org_slug: str, title: str, doc_id: str) -> int:
    """Store raw text as a knowledge entry."""
    save_knowledge_entry(text, org_slug, title, doc_id)
    return 1


def _ingest_csv(file_path: str, org_slug: str, title: str, doc_id: str) -> int:
    """Parse CSV into knowledge entries — one entry per row."""
    entries = 0
    with open(file_path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        rows = list(reader)

    if not rows:
        save_knowledge_entry(f"[Empty CSV file: {title}]", org_slug, title, doc_id)
        return 1

    # Create one entry with the full CSV as structured text
    header = list(rows[0].keys())
    lines = [", ".join(f"{k}: {v}" for k, v in row.items()) for row in rows]
    content = f"Data from {title}:\nColumns: {', '.join(header)}\n\n" + "\n".join(lines)
    save_knowledge_entry(content, org_slug, title, doc_id)
    entries += 1

    # Also create per-row entries for finer-grained search
    for row in rows:
        row_text = ", ".join(f"{k}: {v}" for k, v in row.items())
        save_knowledge_entry(row_text, org_slug, title, doc_id, metadata={"row": True})
        entries += 1

    return entries


def _ingest_xlsx(file_path: str, org_slug: str, title: str, doc_id: str) -> int:
    """Parse XLSX into knowledge entries."""
    from openpyxl import load_workbook

    entries = 0
    wb = load_workbook(file_path, data_only=True)

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # First row as header
        header = [str(c) if c else f"col_{i}" for i, c in enumerate(rows[0])]
        data_rows = rows[1:]

        lines = []
        for row in data_rows:
            pairs = [f"{header[i]}: {row[i]}" for i in range(min(len(header), len(row))) if row[i] is not None]
            if pairs:
                lines.append(", ".join(pairs))

        if lines:
            content = f"Data from {title} (sheet: {sheet}):\nColumns: {', '.join(header)}\n\n" + "\n".join(lines)
            save_knowledge_entry(content, org_slug, f"{title} - {sheet}", doc_id)
            entries += 1

    return entries
